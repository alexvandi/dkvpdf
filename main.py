from flask import Flask, render_template, request, send_file, jsonify
import pdfplumber
import pandas as pd
import re
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

app = Flask(__name__)
UPLOAD_FOLDER = "/tmp"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)


def clean_number(num_str):
    return float(num_str.replace(".", "").replace(",", "."))


@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        try:
            if "pdf" not in request.files:
                return jsonify({"success": False, "error": "Nessun file caricato"}), 400
            
            uploaded_file = request.files["pdf"]
            if not uploaded_file or uploaded_file.filename == "":
                return jsonify({"success": False, "error": "Nessun file selezionato"}), 400
            
            if not uploaded_file.filename.lower().endswith(".pdf"):
                return jsonify({"success": False, "error": "Il file deve essere in formato PDF"}), 400
            
            pdf_name = os.path.splitext(uploaded_file.filename)[0]
            pdf_path = os.path.join(UPLOAD_FOLDER, uploaded_file.filename)
            uploaded_file.save(pdf_path)
            excel_path, excel_name = extract_and_generate_excel(
                pdf_path, pdf_name)
            return jsonify({"success": True, "filename": excel_name})
        except Exception as e:
            return jsonify({"success": False, "error": str(e)}), 500
    return render_template("index.html")


@app.route("/download_excel/<filename>")
def download_excel(filename):
    file_path = os.path.join(UPLOAD_FOLDER, filename)
    if os.path.exists(file_path):
        return send_file(file_path,
                         as_attachment=True,
                         download_name=filename)
    return "Nessun file disponibile", 404


def extract_and_generate_excel(pdf_path, pdf_name):
    rows = []
    current_station = ""
    localita = ""
    current_merce = ""

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text:
                continue
            lines = text.split("\n")

            for line in lines:
                staz_match = re.search(
                    r"Stazione di servizio: (\d+), Italia, (.*?),", line)
                if staz_match:
                    current_station = staz_match.group(1).strip()
                    localita = staz_match.group(2).strip()

                if "Tipo de Merce" in line:
                    parts = line.split("Tipo de Merce")
                    if len(parts) > 1:
                        current_merce = parts[1].strip()

                if re.match(r"\d{2}\.\d{2}\.\d{4}", line):
                    fields = line.split()
                    try:
                        if len(fields) >= 8:
                            data = fields[0]
                            quantita = clean_number(fields[3])
                            unita = fields[4]
                            importo_netto = clean_number(fields[6])
                            if "/" in fields[7]:
                                importo_sconto = clean_number(fields[8])
                            else:
                                importo_sconto = clean_number(fields[7])
                            rows.append([
                                current_station, localita, current_merce, data,
                                quantita, unita, importo_netto, importo_sconto
                            ])
                    except:
                        continue

    df = pd.DataFrame(rows,
                      columns=[
                          "ID Stazione", "Località", "Tipo Merce", "Data",
                          "Quantità", "Unità",
                          "Importo Netto SENZA Sconto (IVA escl.)",
                          "Importo Netto dello Sconto (IVA escl.)"
                      ])

    def crea_riepilogo(campo):
        return df.groupby(["ID Stazione", "Località", "Tipo Merce"
                           ])[campo].sum().reset_index().pivot_table(
                               index=["ID Stazione", "Località"],
                               columns="Tipo Merce",
                               values=campo,
                               fill_value=0).reset_index()

    df_netto = crea_riepilogo("Importo Netto SENZA Sconto (IVA escl.)")
    df_sconto = crea_riepilogo("Importo Netto dello Sconto (IVA escl.)")

    # Create workbook
    excel_filename = f"{pdf_name}.xlsx"
    excel_path = os.path.join(UPLOAD_FOLDER, excel_filename)
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Dati Estesi"

    for r in dataframe_to_rows(df, index=False, header=True):
        ws1.append(r)

    ws2 = wb.create_sheet("Riepilogo")

    def style_tabella(sheet, titolo, df_riga, start_row):
        sheet.merge_cells(start_row=start_row,
                          start_column=1,
                          end_row=start_row,
                          end_column=len(df_riga.columns))
        cell = sheet.cell(row=start_row, column=1)
        cell.value = titolo
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="FFC000")
        cell.alignment = Alignment(horizontal="center")

        for r_idx, r in enumerate(dataframe_to_rows(df_riga,
                                                    index=False,
                                                    header=True),
                                  start=start_row + 1):
            for c_idx, val in enumerate(r, start=1):
                cell = sheet.cell(row=r_idx, column=c_idx, value=val)
                if isinstance(val, (int, float)):
                    cell.number_format = "#,##0.00"
                    cell.alignment = Alignment(horizontal="right")
                    cell.fill = PatternFill(
                        "solid", fgColor="99FF99")  # verde un po' più intenso
                else:
                    cell.alignment = Alignment(horizontal="left")

    style_tabella(ws2, "Conto di ricavo al lordo dello sconto (EURO)",
                  df_netto, 1)
    style_tabella(ws2, "Importo dello sconto (EURO)", df_sconto,
                  len(df_netto) + 5)

    wb.save(excel_path)
    return excel_path, excel_filename


@app.route("/health")
def health_check():
    return "OK", 200

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)
